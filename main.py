import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os


class ExpenseApp(App):
    def build(self):
        self.layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # TextInputs for date range
        self.start_date_input = TextInput(hint_text='Tanggal Mulai (YYYY-MM-DD)', multiline=False, size_hint=(1, None),
                                          height=40, font_size=12, padding=(10, 10))
        self.end_date_input = TextInput(hint_text='Tanggal Akhir (YYYY-MM-DD) opsional', multiline=False,
                                        size_hint=(1, None), height=40, font_size=12, padding=(10, 10))

        self.view_expense_button = Button(text='Lihat Tabel Pengeluaran', size_hint=(1, None), height=40, font_size=12)
        self.view_income_button = Button(text='Lihat Tabel Pemasukan', size_hint=(1, None), height=40, font_size=12)

        # TextInputs for expense and income
        self.amount_input = TextInput(hint_text='Jumlah', multiline=False, size_hint=(1, None), height=40, font_size=12,
                                      padding=(10, 10))
        self.description_input = TextInput(hint_text='Deskripsi', multiline=False, size_hint=(1, None), height=40,
                                           font_size=12, padding=(10, 10))

        self.submit_expense_button = Button(text='Catat Pengeluaran', size_hint=(1, None), height=40, font_size=12)
        self.submit_income_button = Button(text='Catat Pemasukan', size_hint=(1, None), height=40, font_size=12)

        self.record_label = Label(text='', size_hint=(1, None), height=60, font_size=12)

        self.layout.add_widget(self.start_date_input)
        self.layout.add_widget(self.end_date_input)
        self.layout.add_widget(self.view_expense_button)
        self.layout.add_widget(self.view_income_button)
        self.layout.add_widget(self.amount_input)
        self.layout.add_widget(self.description_input)
        self.layout.add_widget(self.submit_expense_button)
        self.layout.add_widget(self.submit_income_button)
        self.layout.add_widget(self.record_label)

        self.submit_expense_button.bind(on_press=self.record_expense)
        self.submit_income_button.bind(on_press=self.record_income)
        self.view_expense_button.bind(on_press=self.show_expense_table)
        self.view_income_button.bind(on_press=self.show_income_table)

        return self.layout

    def record_expense(self, instance):
        amount = self.amount_input.text.strip()
        description = self.description_input.text.strip()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Check for empty fields
        if not amount or not description:
            self.show_warning_popup('Semua field harus diisi!')
            return

        try:
            amount = float(amount)
        except ValueError:
            self.show_warning_popup('Jumlah harus berupa angka!')
            return

        # Save to Excel
        self.save_to_excel(amount, description, current_time, 'pengeluaran.xlsx')

        self.record_label.text = f'Pengeluaran: {amount}\nDeskripsi: {description}\nWaktu: {current_time}'
        self.amount_input.text = ''
        self.description_input.text = ''

    def record_income(self, instance):
        amount = self.amount_input.text.strip()
        description = self.description_input.text.strip()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Check for empty fields
        if not amount or not description:
            self.show_warning_popup('Semua field harus diisi!')
            return

        try:
            amount = float(amount)
        except ValueError:
            self.show_warning_popup('Jumlah harus berupa angka!')
            return

        # Save to Excel
        self.save_to_excel(amount, description, current_time, 'pemasukan.xlsx')

        self.record_label.text = f'Pemasukan: {amount}\nDeskripsi: {description}\nWaktu: {current_time}'
        self.amount_input.text = ''
        self.description_input.text = ''

    def show_warning_popup(self, message):
        content = BoxLayout(orientation='vertical', padding=10)
        content.add_widget(Label(text=message, size_hint=(1, 0.8), font_size=12))
        close_button = Button(text='Tutup', size_hint=(1, 0.2))
        content.add_widget(close_button)

        popup = Popup(title='Peringatan', content=content, size_hint=(0.8, 0.3))
        close_button.bind(on_press=popup.dismiss)
        popup.open()

    def save_to_excel(self, amount, description, current_time, filename):
        file_path = self.get_file_path(filename)

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Jumlah", "Deskripsi", "Waktu"])

        sheet.append([amount, description, current_time])
        workbook.save(file_path)

    def get_file_path(self, filename):
        # Define a fixed path for the Excel file
        directory = os.path.join(os.path.expanduser('~'), 'Documents')
        if not os.path.exists(directory):
            os.makedirs(directory)
        file_path = os.path.join(directory, filename)
        return file_path

    def show_expense_table(self, instance):
        self.show_table('pengeluaran.xlsx', 'Tabel Pengeluaran')

    def show_income_table(self, instance):
        self.show_table('pemasukan.xlsx', 'Tabel Pemasukan')

    def show_table(self, filename, title):
        start_date_str = self.start_date_input.text.strip()
        end_date_str = self.end_date_input.text.strip()

        file_path = self.get_file_path(filename)
        if not os.path.exists(file_path):
            popup = Popup(title='Error', content=Label(text='File tidak ditemukan!'), size_hint=(0.8, 0.8))
            popup.open()
            return

        workbook = load_workbook(file_path)
        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            if row[0] != "Jumlah":  # Skip header
                data.append(row)

        # Sort data by date
        data.sort(key=lambda x: datetime.strptime(x[2], '%Y-%m-%d %H:%M:%S'))

        filtered_data = []
        for row in data:
            row_date = datetime.strptime(row[2], '%Y-%m-%d %H:%M:%S').date()
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                if end_date_str:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                    if start_date <= row_date <= end_date:
                        filtered_data.append(row)
                else:
                    if row_date == start_date:
                        filtered_data.append(row)
            else:
                filtered_data = data

        table_layout = GridLayout(cols=3, size_hint_y=None)
        table_layout.bind(minimum_height=table_layout.setter('height'))

        # Add header
        table_layout.add_widget(Label(text="Jumlah", font_size=12, bold=True))
        table_layout.add_widget(Label(text="Deskripsi", font_size=12, bold=True))
        table_layout.add_widget(Label(text="Waktu", font_size=12, bold=True))

        total_amount = 0
        for row in filtered_data:
            for cell in row:
                table_layout.add_widget(Label(text=str(cell), font_size=12))
            total_amount += float(row[0])

        scroll_view = ScrollView(size_hint=(1, None), height=400)
        scroll_view.add_widget(table_layout)

        popup_layout = BoxLayout(orientation='vertical')
        popup_layout.add_widget(scroll_view)

        total_label = Label(text=f'Total: {total_amount}', size_hint=(1, None), height=40, font_size=12)
        popup_layout.add_widget(total_label)

        close_button = Button(text='Tutup', size_hint=(1, None), height=40)
        close_button.bind(on_press=lambda x: popup.dismiss())
        popup_layout.add_widget(close_button)

        popup = Popup(title=title, content=popup_layout, size_hint=(0.9, 0.9))
        popup.open()


if __name__ == '__main__':
    ExpenseApp().run()
